# imports
from tkinter import *
from tkinter import messagebox
from tkinter import ttk

from PIL import Image, ImageTk

import openpyxl
from openpyxl import *
from openpyxl.styles import Font, Color

import os.path
from os import path


# ====== initialisation=========
root = Tk()
root.title("Student Registration Form")
root.geometry("1350x700+0+0")
root.config(bg="teal")


if path.exists('data.xlsx'):
    Wb = load_workbook('data.xlsx')
else:
    Wb = openpyxl.Workbook()
    Wb.save('data.xlsx')

sheet = Wb.active

# ======== functions============


def excel():
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 60
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 50
    sheet.column_dimensions['G'].width = 50
    sheet.column_dimensions['H'].width = 50
    sheet.column_dimensions['I'].width = 30
    sheet.column_dimensions['J'].width = 30
    sheet.column_dimensions['K'].width = 30
    sheet.column_dimensions['L'].width = 30
    sheet.column_dimensions['M'].width = 30
    sheet.column_dimensions['N'].width = 30
    sheet.column_dimensions['O'].width = 30

    sheet.cell(row=1, column=1).value = "Student Name"
    sheet.cell(row=1, column=2).value = "Father's Name"
    sheet.cell(row=1, column=3).value = "Mother Name"
    sheet.cell(row=1, column=4).value = "Email"
    sheet.cell(row=1, column=5).value = "Gender"
    sheet.cell(row=1, column=6).value = "Contact Number"
    sheet.cell(row=1, column=7).value = "Present Address"
    sheet.cell(row=1, column=8).value = "Permanent Address"
    sheet.cell(row=1, column=9).value = "X Marks"
    sheet.cell(row=1, column=10).value = "XII Marks"
    sheet.cell(row=1, column=11).value = "Board"
    sheet.cell(row=1, column=12).value = "Physics Marks"
    sheet.cell(row=1, column=13).value = "Chemistry Marks"
    sheet.cell(row=1, column=14).value = "Maths Marks"
    sheet.cell(row=1, column=15).value = "Percentage Marks"

    a1 = sheet['A1']
    b1 = sheet['B1']
    c1 = sheet['C1']
    d1 = sheet['D1']
    e1 = sheet['E1']
    f1 = sheet['F1']
    g1 = sheet['G1']
    h1 = sheet['H1']
    i1 = sheet['I1']
    j1 = sheet['J1']
    k1 = sheet['K1']
    l1 = sheet['L1']
    m1 = sheet['M1']
    n1 = sheet['N1']
    o1 = sheet['O1']

    a1.font = Font(bold=True)
    b1.font = Font(bold=True)
    c1.font = Font(bold=True)
    d1.font = Font(bold=True)
    e1.font = Font(bold=True)
    f1.font = Font(bold=True)
    g1.font = Font(bold=True)
    h1.font = Font(bold=True)
    i1.font = Font(bold=True)
    j1.font = Font(bold=True)
    k1.font = Font(bold=True)
    l1.font = Font(bold=True)
    m1.font = Font(bold=True)
    n1.font = Font(bold=True)
    o1.font = Font(bold=True)


def removeData():
    ck = chk_var.get()
    if(ck != 0):
        student_name.delete(0, END)
        fname.delete(0, END)
        mname.delete(0, END)
        mail.delete(0, END)
        phone.delete(0, END)
        present_address.delete(0, END)
        permanent_address.delete(0, END)
        ten.delete(0, END)
        twelve.delete(0, END)
        phy.delete(0, END)
        chem.delete(0, END)
        math.delete(0, END)
        cmb_board.delete(0, END)
        gender.set(None)


def getData():
    chk = chk_var.get()
    lst = []
    content = len(studentName.get())
    lst.append(content)
    content = len(fatherName.get())
    lst.append(content)
    content = len(motherName.get())
    lst.append(content)
    content = len(email.get())
    lst.append(content)
    content = gender.get()
    lst.append(content)
    content = len(phoneNumber.get())
    lst.append(content)
    content = len(presentAddress.get())
    lst.append(content)
    content = len(permanentAddress.get())
    lst.append(content)
    content = len(tenMarks.get())
    lst.append(content)
    content = len(twelveMarks.get())
    lst.append(content)
    content = len(cmb_board.get())
    lst.append(content)
    content = len(phyMarks.get())
    lst.append(content)
    content = len(chemMarks.get())
    lst.append(content)
    content = len(mathMarks.get())
    lst.append(content)
    # print(lst)

    if (chk == 0):
        messagebox.showerror('Error', 'Please check the checkbutton')

    if (0 in lst):
        messagebox.showerror('Error', 'Please fill all the fields')

    else:
        currentRow = sheet.max_row
        #currentColumn = sheet.max_column

        sheet.cell(row=currentRow+1, column=1).value = studentName.get()
        sheet.cell(row=currentRow+1, column=2).value = fatherName.get()
        sheet.cell(row=currentRow+1, column=3).value = motherName.get()
        sheet.cell(row=currentRow+1, column=4).value = email.get()
        Gender_res = gender.get()

        if Gender_res == 1:
            sheet.cell(row=currentRow+1, column=5).value = "Male"
        elif Gender_res == 2:
            sheet.cell(row=currentRow+1, column=5).value = "Female"
        else:
            sheet.cell(row=currentRow+1, column=5).value = "Other"

        #sheet.cell(row=currentRow+1, column=5).value = gender.get()
        sheet.cell(row=currentRow+1, column=6).value = phoneNumber.get()
        sheet.cell(row=currentRow+1, column=7).value = presentAddress.get()
        sheet.cell(row=currentRow+1, column=8).value = permanentAddress.get()
        sheet.cell(row=currentRow+1, column=9).value = tenMarks.get()
        sheet.cell(row=currentRow+1, column=10).value = twelveMarks.get()
        sheet.cell(row=currentRow+1, column=11).value = cmb_board.get()
        sheet.cell(row=currentRow+1, column=12).value = phyMarks.get()
        sheet.cell(row=currentRow+1, column=13).value = chemMarks.get()
        sheet.cell(row=currentRow+1, column=14).value = mathMarks.get()
        Percentage = round(
            (int(phyMarks.get()) + int(chemMarks.get()) + int(mathMarks.get()))/3)
        sheet.cell(row=currentRow+1, column=15).value = Percentage

        Wb.save('data.xlsx')
        Wb.close()
        removeData()


def show():
    excel()
    getData()


# ==================== Declarations =========================
studentName = StringVar()
fatherName = StringVar()
motherName = StringVar()
email = StringVar()
gender = IntVar()
phoneNumber = StringVar()
presentAddress = StringVar()
permanentAddress = StringVar()
tenMarks = StringVar()
twelveMarks = StringVar()
otherBoard = StringVar()
phyMarks = StringVar()
chemMarks = StringVar()
mathMarks = StringVar()
chk_var = IntVar()

# ======================== Form ==============================

# BG IMAGE #
bg_image = Image.open(r"images\stcet.png")
bg = ImageTk.PhotoImage(bg_image)
bg_label = Label(image=bg)
bg_label.place(x=0, y=0, relwidth=1, relheight=1)

#COLLEGE LOGO #
logo_image = Image.open(r"images\logo.png")
logo = ImageTk.PhotoImage(logo_image)
logo_label = Label(image=logo)
logo_label.place(x=140, y=100, width=290, height=480)

#  REGISTRATION FRAME #

# we are making the frame in the root window
frame1 = Frame(root, bg="white")
frame1.place(x=430, y=100, width=730, height=480)

# title
title = Label(frame1, text="PLEASE ENTER YOUR DETAILS HERE", font=(
    "times new roman", 18, "bold"), bg="white", fg="brown")
title.place(x=70, y=5)

# STUDENT NAME #
student_Label = Label(frame1, text="NAME  ", font=(
    "verdana", 10, "bold"), bg="white", fg="black")
student_Label.place(x=10, y=50)

student_name = Entry(frame1, font=("verdana", 10),
                     textvariable=studentName, bg="lightgray")
student_name.place(x=150, y=50)

#PARENT'S NAME #
fname_label = Label(frame1, text="FATHER'S NAME  ", font=(
    "verdana", 10, "bold"), bg="white", fg="black")
fname_label.place(x=10, y=80)

fname = Entry(frame1, font=("verdana", 10),
              bg="lightgray", textvariable=fatherName)
fname.place(x=150, y=80)

mname_label = Label(frame1, text="MOTHER'S NAME  ", font=(
    "verdana", 10, "bold"), bg="white", fg="black")
mname_label.place(x=360, y=80)

mname = Entry(frame1, font=("verdana", 10),
              bg="lightgray", textvariable=motherName)
mname.place(x=500, y=80)


# MAIL ID #
mail_label = Label(frame1, text="EMAIL ID  ", font=(
    "verdana", 10, "bold"), bg="white", fg="black")
mail_label.place(x=10, y=110)

mail = Entry(frame1, font=("verdana", 10),
             bg="lightgray", textvariable=email, width=30)
mail.place(x=150, y=110)


# SEX #
sex_label = Label(frame1, text="SEX  ", font=(
    "verdana", 10, "bold"), bg="white", fg="black")
sex_label.place(x=10, y=140)

# RADIOBUTTON
Radiobutton(frame1, text="MALE", value=1,
            variable=gender).place(x=150, y=140)

Radiobutton(frame1, text="FEMALE", value=2,
            variable=gender).place(x=213, y=140)

Radiobutton(frame1, text="OTHER", value=3,
            variable=gender).place(x=288, y=140)


# CONTACT NUMBER #
phone_label = Label(frame1, text="PHONE NUMBER  ", font=(
    "verdana", 10, "bold"), bg="white", fg="black")
phone_label.place(x=10, y=170)

phone = Entry(frame1, font=("verdana", 10),
              bg="lightgray", textvariable=phoneNumber)
phone.place(x=150, y=170)

# ADDRESS #
present_address_label = Label(frame1, text="PRESENT ADDRESS  ", font=(
    "verdana", 10, "bold"), bg="white", fg="black")
present_address_label.place(x=10, y=200)

present_address = Entry(frame1, font=("verdana", 10),
                        bg="lightgray", textvariable=presentAddress)
present_address.place(x=150, y=200)

permanent_address_label = Label(frame1, text="PERMANENT ADDRESS  ", font=(
    "verdana", 10, "bold"), bg="white", fg="black")
permanent_address_label.place(x=320, y=200)

permanent_address = Entry(frame1, font=("verdana", 10),
                          bg="lightgray", textvariable=permanentAddress)
permanent_address.place(x=500, y=200)


# CLASS 10 #
ten_label = Label(frame1, text="X BOARD MARKS", font=(
    "verdana", 10, "bold"), bg="white", fg="black")
ten_label.place(x=10, y=230)

ten = Entry(frame1, font=("verdana", 10),
            bg="lightgray", textvariable=tenMarks)
ten.place(x=150, y=230)


# CLASS 12 #
twelve_label = Label(frame1, text="XII BOARD MARKS", font=(
    "verdana", 10, "bold"), bg="white", fg="black")
twelve_label.place(x=350, y=230)

twelve = Entry(frame1, font=("verdana", 10),
               bg="lightgray", textvariable=twelveMarks)
twelve.place(x=500, y=230)


# BOARD NAME #
board_label = Label(frame1, text="BOARD", font=(
    "verdana", 10, "bold"), bg="white", fg="black")
board_label.place(x=10, y=260)

cmb_board = ttk.Combobox(frame1, font=(
    "verdane", 10, "bold"), values=("ISC", "STATE BOARD", "CBSE", "OTHER"))
cmb_board.place(x=150, y=260)


# PCM MARKS #
phy_label = Label(frame1, text="MARKS IN PHY", font=(
    "verdana", 10, "bold"), bg="white", fg="black")
phy_label.place(x=10, y=290)

phy = Entry(frame1, font=("verdana", 10),
            bg="lightgray", textvariable=phyMarks)
phy.place(x=150, y=290)

chem_label = Label(frame1, text="MARKS IN CHEM", font=(
    "verdana", 10, "bold"), bg="white", fg="black")
chem_label.place(x=10, y=320)

chem = Entry(frame1, font=("verdana", 10),
             bg="lightgray", textvariable=chemMarks)
chem.place(x=150, y=320)

math_label = Label(frame1, text="MARKS IN MATH", font=(
    "verdana", 10, "bold"), bg="white", fg="black")
math_label.place(x=10, y=350)

math = Entry(frame1, font=("verdana", 10),
             bg="lightgray", textvariable=mathMarks)
math.place(x=150, y=350)

# checkbox
chk_1 = Checkbutton(frame1,
                    text="I do hereby declare that all the information given above is true to the best of my knowledge and belief.",
                    bg="LightBlue1", activebackground="#80c1ff", font=("bold", 11), bd=15, variable=chk_var)
chk_1.place(x=10, y=380)

# SUBMIT BUTTON
Button(frame1, font=("verdana", 12), text="Submit", bg="#15A01F", activebackground="#75F87E",
       command=show).place(x=260, y=440)

# RESET BUTTON
Button(frame1, font=("verdana", 12), text="Reset", bg="#D74627", activebackground="#F96259",
       command=removeData).place(x=360, y=440)


root.mainloop()
